"""
export-viberate.py
Exports viberate_metrics_daily to a formatted Excel workbook.
Run from backend/ folder:
  python ../export-viberate.py
Or place anywhere and adjust DB connection.
"""

import os, sys
import psycopg2
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── DB connection from environment ──────────────────────────────────────────
DB_URL = os.environ.get("DATABASE_URL") or os.environ.get("DIRECT_URL")
if not DB_URL:
    print("ERROR: Set DATABASE_URL in your environment or .env file")
    sys.exit(1)

# Strip ?schema=... suffix that Prisma sometimes appends
if "?" in DB_URL:
    DB_URL = DB_URL.split("?")[0]

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# ── Styles ───────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F2937")   # dark
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHDR_FILL   = PatternFill("solid", start_color="374151")
SUBHDR_FONT   = Font(name="Arial", bold=True, color="F9FAFB", size=9)
DATA_FONT     = Font(name="Arial", size=9)
ALT_FILL      = PatternFill("solid", start_color="F9FAFB")
BORDER_SIDE   = Side(style="thin", color="E5E7EB")
THIN_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE, bottom=BORDER_SIDE)

PLATFORM_COLORS = {
    "spotify":   "1DB954",
    "youtube":   "FF0000",
    "instagram": "E1306C",
    "facebook":  "1877F2",
    "tiktok":    "010101",
}

def style_header(cell, fill_color=None):
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", start_color=fill_color or "1F2937")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

def style_data(cell, alt=False):
    cell.font = DATA_FONT
    if alt:
        cell.fill = ALT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")

def num_fmt(cell, fmt):
    cell.number_format = fmt

# ── Fetch data ───────────────────────────────────────────────────────────────
cur.execute("""
    SELECT
        a."artistName",
        v."metricName",
        v."date",
        v."diffValue",
        v."totalValue",
        v."apiVersion",
        v."fetchedAt"
    FROM viberate_metrics_daily v
    JOIN artists a ON a.id = v."artistId"
    ORDER BY a."artistName", v."metricName", v."date"
""")
rows = cur.fetchall()

cur.execute('SELECT DISTINCT "artistName" FROM artists WHERE "viberateSlug" IS NOT NULL ORDER BY 1')
artists = [r[0] for r in cur.fetchall()]

cur.execute('SELECT DISTINCT "metricName" FROM viberate_metrics_daily ORDER BY 1')
metrics = [r[0] for r in cur.fetchall()]

conn.close()

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Sheet 1: Summary (all artists, latest total per metric) ──────────────────
ws_sum = wb.create_sheet("Summary")

# Title
ws_sum.merge_cells("A1:H1")
ws_sum["A1"] = f"Viberate Artist Intelligence — Summary"
ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="111827")
ws_sum["A1"].alignment = Alignment(horizontal="center")
ws_sum.row_dimensions[1].height = 28

ws_sum.merge_cells("A2:H2")
ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
ws_sum["A2"].font = Font(name="Arial", size=9, color="6B7280")
ws_sum["A2"].alignment = Alignment(horizontal="center")

# Headers row 4
summary_headers = [
    "Artist", "Platform", "Metric",
    "Latest Total", "Latest Date",
    "30d Change", "90d Change", "Data Points"
]
for col, h in enumerate(summary_headers, 1):
    c = ws_sum.cell(row=4, column=col, value=h)
    style_header(c)
ws_sum.row_dimensions[4].height = 20

# Build summary rows — get latest total per artist+metric
from collections import defaultdict
data_by_artist_metric = defaultdict(list)
for artist, metric, date, diff, total, api, fetched in rows:
    data_by_artist_metric[(artist, metric)].append((date, diff, total))

summary_row = 5
for artist in artists:
    for metric in metrics:
        key = (artist, metric)
        if key not in data_by_artist_metric:
            continue
        points = sorted(data_by_artist_metric[key], key=lambda x: x[0])

        latest = next((p for p in reversed(points) if p[2] is not None), None)
        if not latest:
            continue

        latest_total = latest[2]
        latest_date = latest[0]

        # 30d change: sum diffs in last 30 days
        from datetime import timedelta
        cutoff_30 = latest_date - timedelta(days=30)
        cutoff_90 = latest_date - timedelta(days=90)
        change_30 = sum(p[1] or 0 for p in points if p[0] >= cutoff_30 and p[1] != 0)
        change_90 = sum(p[1] or 0 for p in points if p[0] >= cutoff_90 and p[1] != 0)

        platform = metric.split("_")[0]
        alt = (summary_row % 2 == 0)

        vals = [
            artist, platform, metric,
            latest_total, latest_date.strftime("%Y-%m-%d"),
            change_30, change_90, len(points)
        ]
        for col, val in enumerate(vals, 1):
            c = ws_sum.cell(row=summary_row, column=col, value=val)
            style_data(c, alt)
            if col in (4, 6, 7):
                c.number_format = "#,##0"
            if col == 2:
                color = PLATFORM_COLORS.get(platform, "6B7280")
                c.font = Font(name="Arial", size=9, bold=True, color=color)

        summary_row += 1

# Column widths for summary
for col, w in zip("ABCDEFGH", [22, 12, 30, 16, 14, 14, 14, 12]):
    ws_sum.column_dimensions[col].width = w

ws_sum.freeze_panes = "A5"

# ── Sheet per artist: time series ─────────────────────────────────────────────
for artist in artists:
    safe_name = artist[:31]  # Excel sheet name limit
    ws = wb.create_sheet(safe_name)

    # Get all metrics and dates for this artist
    artist_rows = [(m, d, diff, total) for a, m, d, diff, total, *_ in rows if a == artist]
    if not artist_rows:
        continue

    artist_metrics = sorted(set(r[0] for r in artist_rows))
    all_dates = sorted(set(r[1] for r in artist_rows))

    # Build lookup: metric → date → (diff, total)
    lookup = defaultdict(dict)
    for m, d, diff, total in artist_rows:
        lookup[m][d] = (diff, total)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(artist_metrics)*2+1)}1")
    ws["A1"] = f"{artist} — Viberate Daily Metrics"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="111827")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    # Header row 3: Date | metric1_diff | metric1_total | metric2_diff ...
    ws.cell(row=3, column=1, value="Date")
    style_header(ws.cell(row=3, column=1))

    col = 2
    metric_col_map = {}
    for metric in artist_metrics:
        platform = metric.split("_")[0]
        color = PLATFORM_COLORS.get(platform, "374151")

        c_diff = ws.cell(row=3, column=col, value=f"{metric}\n(diff)")
        c_diff.font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
        c_diff.fill = PatternFill("solid", start_color=color)
        c_diff.alignment = Alignment(horizontal="center", wrap_text=True)
        c_diff.border = THIN_BORDER

        c_total = ws.cell(row=3, column=col+1, value=f"{metric}\n(total)")
        c_total.font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
        c_total.fill = PatternFill("solid", start_color=color)
        c_total.fill = PatternFill("solid", start_color="4B5563")
        c_total.alignment = Alignment(horizontal="center", wrap_text=True)
        c_total.border = THIN_BORDER

        metric_col_map[metric] = col
        col += 2

    ws.row_dimensions[3].height = 32

    # Data rows
    for row_idx, date in enumerate(all_dates):
        data_row = row_idx + 4
        alt = (row_idx % 2 == 0)

        c = ws.cell(row=data_row, column=1, value=date.strftime("%Y-%m-%d"))
        style_data(c, alt)

        for metric in artist_metrics:
            mc = metric_col_map[metric]
            diff_val, total_val = lookup[metric].get(date, (None, None))

            c_diff = ws.cell(row=data_row, column=mc, value=diff_val)
            c_total = ws.cell(row=data_row, column=mc+1, value=total_val)

            style_data(c_diff, alt)
            style_data(c_total, alt)

            if diff_val is not None:
                c_diff.number_format = "#,##0"
                # Red for negative, green for positive diff
                if diff_val < 0:
                    c_diff.font = Font(name="Arial", size=9, color="DC2626")
                elif diff_val > 0:
                    c_diff.font = Font(name="Arial", size=9, color="059669")

            if total_val is not None:
                c_total.number_format = "#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 13
    for metric in artist_metrics:
        mc = metric_col_map[metric]
        ws.column_dimensions[get_column_letter(mc)].width = 16
        ws.column_dimensions[get_column_letter(mc+1)].width = 16

    ws.freeze_panes = "B4"

# ── Sheet: Metrics Legend ──────────────────────────────────────────────────────
ws_leg = wb.create_sheet("Legend")
ws_leg.merge_cells("A1:C1")
ws_leg["A1"] = "Metrics Reference"
ws_leg["A1"].font = Font(name="Arial", bold=True, size=13)
ws_leg["A1"].alignment = Alignment(horizontal="center")

legend_data = [
    ("Metric Name", "Platform", "Notes"),
    ("spotify_followers", "Spotify", "Total followers count"),
    ("spotify_listeners", "Spotify", "Monthly listeners"),
    ("spotify_streams",   "Spotify", "Cumulative streams"),
    ("spotify_popularity","Spotify", "0-100 score, updated infrequently"),
    ("youtube_subscribers","YouTube","Total channel subscribers"),
    ("youtube_views",     "YouTube", "Cumulative video views"),
    ("youtube_channel_views","YouTube","Channel-level view count, updates weekly"),
    ("youtube_likes",     "YouTube", "Cumulative likes across videos"),
    ("instagram_followers","Instagram","Total followers"),
    ("instagram_likes",   "Instagram","Daily likes (no running total from Viberate)"),
    ("instagram_comments","Instagram","Daily comments (no running total from Viberate)"),
    ("facebook_followers","Facebook","Total page followers"),
    ("tiktok_followers",  "TikTok",  "Total followers"),
    ("tiktok_channel_likes","TikTok","Total channel likes"),
    ("tiktok_comments",   "TikTok",  "Daily comments (no running total)"),
    ("tiktok_views",      "TikTok",  "Daily views (no running total)"),
]

for row_idx, (a, b, c) in enumerate(legend_data):
    r = row_idx + 3
    is_header = row_idx == 0
    for col_idx, val in enumerate([a, b, c], 1):
        cell = ws_leg.cell(row=r, column=col_idx, value=val)
        if is_header:
            style_header(cell)
        else:
            style_data(cell, row_idx % 2 == 0)
            if col_idx == 2:
                platform = b.lower()
                color = PLATFORM_COLORS.get(platform, "6B7280")
                cell.font = Font(name="Arial", size=9, bold=True, color=color)

ws_leg.column_dimensions["A"].width = 28
ws_leg.column_dimensions["B"].width = 14
ws_leg.column_dimensions["C"].width = 45
ws_leg.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = f"viberate_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(out_path)
print(f"✓ Saved: {out_path}")
print(f"  Artists: {len(artists)}")
print(f"  Metrics: {len(metrics)}")
print(f"  Total rows: {len(rows)}")
